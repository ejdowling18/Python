import time
from bs4 import BeautifulSoup
import glob
import csv
import requests
import lxml
import pandas
from xlsxwriter.workbook import Workbook
import pandas as pd
from openpyxl import load_workbook
import sys
import numpy as np
def get_weeks_available(wb):
    total_weeks = 0
    for i in range(1, 17, 1):
        if 'week ' + str(i) in wb.sheetnames:
            if i > total_weeks:
                total_weeks = i
    return total_weeks


wb = load_workbook('FootballAgg.xlsx')
total_weeks = get_weeks_available(wb)

df = pd.DataFrame({'Week': [], 'Year': [], 'GID': [], 'Name': [], 'Pos': [], 'Team': [], 'h/a': [], 'Oppt': [],
                       'FD points': [], 'FD salary': []})

# load all existing data from FootballAgg.xlsx
for i in range(1, total_weeks + 1):
    df_tmp = pd.read_excel('FootballAgg.xlsx', 'week ' + str(i))
    if not (df_tmp.empty):
        df = pd.concat([df, df_tmp])

# Getting a list of each position
positions = list(set(df['Pos']))
for position in positions:
    # Filter by position
    df_position = df[df.Pos == position]
    name_position_data = {'Name': [], 'GID': []}
    # populate data with the amount of keys we need, for the number of weeks
    for i in range(1, total_weeks + 1):
        name_position_data['Week ' + str(i) + ' FD points'] = []
        name_position_data['Week ' + str(i) + ' FD salary'] = []
    # make new dataframe
    df_name_position = pd.DataFrame(name_position_data)
    # list of all names
    names = list(set(df_position['Name']))
    for name in names:
        # filter by name
        df_tmp = df_position[df_position.Name == name]
        df_tmp = df_tmp.reset_index(drop=True)
        tmp_data = {'Name': [name], 'GID': [df_tmp.at[0, 'GID']]}
        for i in range(1, total_weeks + 1):
            # filter by week
            df_week = df_tmp[df_tmp.Week == i]
            df_week = df_week.reset_index(drop=True)
            # making sure the player played this week
            if not df_week.empty:
                # grabbing and setting the values for the position->name->week that we're on
                tmp_data['Week ' + str(i) + ' FD points'] = [df_week.at[0, 'FD points']]
                tmp_data['Week ' + str(i) + ' FD salary'] = [df_week.at[0, 'FD salary']]
        df_tmp = pd.DataFrame(tmp_data)
        df_name_position = pd.concat([df_name_position, df_tmp])

            #df_w = grouped['FD points']
            #df_w["W"+str(i)] = grouped['Week']

    book = load_workbook('FootballTrends.xlsx')
    writer = pandas.ExcelWriter('FootballTrends.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df_name_position.to_excel(writer, sheet_name=position, index=False)

    writer.save()