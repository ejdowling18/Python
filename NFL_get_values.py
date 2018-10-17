import time
from bs4 import BeautifulSoup
import glob
import csv
import requests
import lxml
import pandas
from xlsxwriter.workbook import Workbook
import pandas as pd
from openpyxl import load_workbook
import sys
if sys.version_info[0] < 3:
    from StringIO import StringIO
else:
    from io import StringIO

from selenium import webdriver


#driver = webdriver.Chrome()
#driver.get("https://rotogrinders.com/lineuphq/mlb?site=fanduel")
#time.sleep(10)
##button = driver.find_element_by_id('build-button')
#button.click()
#time.sleep(5)
#button = driver.find_element_by_xpath('//button[@class="button button--full"]')
#button.click()
proxies = {'http': 'http://ny-i-bluecoat.dslocal.com:8080',
    'https': 'https://ny-i-bluecoat.dslocal.com:8080'}
# navigating to the URL specified in the r row of our csv
for i in range(1,17,1):
    i = str(i)
    page = requests.get('http://rotoguru1.com/cgi-bin/fyday.pl?week='+i+'&game=fd&scsv=1', proxies=proxies)

# reading in the pages content
    c = page.content
# parsing the page into a BeautifulSoup object, using the lxml module
    soup = BeautifulSoup(c, features='lxml')
    text_container =  soup.select("pre")[0].text

##text_container = StringIO(text_container)
##print(text_container)

    df = pd.DataFrame({'Week': [], 'Year': [], 'GID': [], 'Name': [], 'Pos': [], 'Team': [], 'h/a': [], 'Oppt': [],
                   'FD points': [], 'FD salary': []})

    for text in text_container.split('\n')[1:]:
        if ';' in text:
            text_list = text.split(';')
            df_tmp = pd.DataFrame({'Week': [text_list[0]], 'Year': [text_list[1]], 'GID': [text_list[2]],
                                'Name': [text_list[3]], 'Pos': [text_list[4]], 'Team': [text_list[5]],
                               'h/a': [text_list[6]], 'Oppt': [text_list[7]], 'FD points': [text_list[8]],
                               'FD salary': [text_list[9]]})
            df = pd.concat([df, df_tmp])
            df = df.sort_values(by=['GID'])
    #print(df)

#text_container = text_container.split("\n")
#for text in text_container :
 #   print(text)
#book = load_workbook('Football.xlsx')
#writer = pd.ExcelWriter('Football.xlsx')
##writer.book = book
#writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    if not(df.empty):
        book = load_workbook('FootballAgg.xlsx')
        writer = pandas.ExcelWriter('FootballAgg.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        df.to_excel(writer, sheet_name='week '+i, index = False)

        writer.save()

#df.to_excel(writer, sheet_name='Week 2', index= False)

#writer.save()
