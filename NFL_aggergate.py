import time
from bs4 import BeautifulSoup
import glob
import csv
import requests
import lxml
import pandas
from xlsxwriter.workbook import Workbook
import pandas as pd
from openpyxl import load_workbook
import sys
import numpy as np
wb = load_workbook('FootballAgg.xlsx')
df = pd.DataFrame({'Week': [], 'Year': [], 'GID': [], 'Name': [], 'Pos': [], 'Team': [], 'h/a': [], 'Oppt': [],
                       'FD points': [], 'FD salary': []})
for i in range(1,17,1):
    i = str(i)
    if 'week '+i in wb.sheetnames:
        df_tmp = pd.read_excel('FootballAgg.xlsx', 'week '+i)
        if not(df_tmp.empty):
            df = pd.concat([df, df_tmp])
            positions = list(set(df['Pos']))
            for position in positions:
                df_position = df[df.Pos == position]
                book = load_workbook('FootballAgg.xlsx')
                writer = pandas.ExcelWriter('FootballAgg.xlsx', engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                df_position.to_excel(writer, sheet_name=position, index=False)

                writer.save()
for i in range(1,17,1):
    i = str(i)
    if 'week '+i in wb.sheetnames:
        df_tmp = pd.read_excel('FootballAgg.xlsx', 'week '+i)
        if not(df_tmp.empty):
            df = pd.concat([df, df_tmp])
            positions = list(set(df['Pos']))
            for position in positions:
                df_position = df[df.Pos == position]
                grouped = df_position.groupby(['GID', 'Name', 'Team'])
                df_avg = grouped['FD points', 'FD salary'].agg(np.mean).reset_index()
                df_avg ['Salary/Points'] = df_avg['FD salary'] / df_avg['FD points']
                book = load_workbook('FootballAgg.xlsx')
                writer = pandas.ExcelWriter('FootballAgg.xlsx', engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df_avg = df_avg.sort_values(by=['Salary/Points'])
                df_avg.to_excel(writer, sheet_name=position+'AVG', index=False)

                writer.save()
for i in range(1, 17, 1):
    i = str(i)
    if 'week ' + i in wb.sheetnames:
        df_tmp = pd.read_excel('FootballAgg.xlsx', 'week ' + i)
        if not (df_tmp.empty):
            df = pd.concat([df, df_tmp])
            positions = list(set(df['Pos']))
            for position in positions:
                df_position = df[df.Pos == position]
                grouped = df_position.groupby(['Oppt'])
                df_avg = grouped['FD points'].agg(np.mean).reset_index()
                book = load_workbook('FootballAgg.xlsx')
                writer = pandas.ExcelWriter('FootballAgg.xlsx', engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df_avg = df_avg.sort_values(by=['FD points'], ascending=False)
                df_avg.to_excel(writer, sheet_name=position + 'OPP', index=False)

                writer.save()

#grouped = df.groupby(['GID', 'Name'])
#grouped = grouped['FD points'].agg(np.mean)
#print (grouped)
